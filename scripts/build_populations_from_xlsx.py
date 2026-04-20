#!/usr/bin/env python3
"""
build_populations_from_xlsx.py — Extract ONS mid-2023 populations
from the official xlsx into a JSON lookup usable by the frontend.

Why this exists: the previous data/uk/ons/gss_population_mid2023.json
only contained 19 local authorities (hand-sampled). The xlsx has 345.
The frontend's Tier-C procedural enrichment needs all of them so that
clicks on London boroughs, Met districts, shire counties, etc. can
show per-capita metrics.

Source: ONS Estimates of the Population for England and Wales, mid-2023.
URL: https://www.ons.gov.uk/peoplepopulationandcommunity/populationandmigration/populationestimates/datasets/estimatesofthepopulationforenglandandwales
Downloaded file: data/uk/ons/mye23tablesew.xlsx (sheet: MYE2 - Persons)

Run:    py scripts/build_populations_from_xlsx.py
Output: data/uk/ons/gss_population_mid2023.json
"""
import json
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data/uk/ons/mye23tablesew.xlsx"
OUT = ROOT / "data/uk/ons/gss_population_mid2023.json"


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["MYE2 - Persons"]
    # Header row is row 8 ("Code", "Name", "Geography", "All ages", ...).
    # Data rows start at row 9.
    all_rows = {}
    for r in range(9, ws.max_row + 1):
        code = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        geog = ws.cell(r, 3).value
        pop = ws.cell(r, 4).value
        if not (code and name and pop):
            continue
        all_rows[code] = {"name": name, "population": int(pop), "geography": geog}

    # Keep only LA-level rows (exclude countries + regions, which we separate
    # as reference aggregates).
    la_level = {
        k: v for k, v in all_rows.items()
        if v["geography"] not in ("Country", "Region")
    }

    out = {
        "_meta": {
            "source": "ONS Local Authority Population Estimates",
            "edition": "Mid-2023",
            "release_date": "2024-07-15",
            "sheet_used": "MYE2 - Persons",
            "column_used": "All ages",
            "downloaded_from": (
                "https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/"
                "populationandmigration/populationestimates/datasets/"
                "estimatesofthepopulationforenglandandwales/"
                "mid20232023localauthorityboundarieseditionofthisdataset/"
                "mye23tablesew.xlsx"
            ),
            "generated_by": "scripts/build_populations_from_xlsx.py",
            "note": (
                f"England + Wales only ({len(la_level)} LAs). "
                "Police (E23xxx) and Fire (E31xxx) entities are NOT in this file — "
                "they are service areas, not geographic local authorities."
            ),
        },
        "reference_aggregates": {
            "E92000001": all_rows.get("E92000001", {}),
            "K04000001": all_rows.get("K04000001", {}),
        },
        "populations": la_level,
    }

    OUT.write_text(json.dumps(out, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Wrote {len(la_level)} LA populations to {OUT.relative_to(ROOT)}")
    # Quick sanity: verify a few well-known LAs are present.
    by_name = {v["name"].lower(): v for v in la_level.values()}
    for check in ["tower hamlets", "kensington and chelsea", "westminster",
                  "birmingham", "manchester", "cornwall", "kent"]:
        hit = by_name.get(check)
        print(f"  {check}: {hit['population'] if hit else 'NOT FOUND'}")


if __name__ == "__main__":
    main()
